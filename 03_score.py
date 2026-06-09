# ==========================================================================
# Phase 3: 汉字结构测评系统 —— 准确率统计阶段
# ==========================================================================
# 输入：
# 1. eval_results/clean_{model}.jsonl
# 2. answer_action.json
#
# 输出：
# 1. eval_results/final_{model}.jsonl
# 2. eval_results/summary_scores.xlsx
#
# 本版修改：
# - 评分完成后，把各模型总分和 add/replace/delete/keep/unknown 分项得分写入 Excel。
# ==========================================================================

import json
import os
from pathlib import Path
from typing import Any, Dict, List, Tuple

from tqdm import tqdm


# ==========================================================================
# 1. 基础配置区
# ==========================================================================

MODEL_LIST = [
    "claude-opus-4-6",
    "gemini-3-flash-preview",
    "glm-4.7",
    "gemini-3-pro-preview",
    "gpt-5.4",
    "deepseek-v3.2",
    "kimi-k2.5"
]

ACTION_FILE = "answer_action.json"
OUTPUT_DIR = "eval_results"
EXCEL_SUMMARY_FILE = os.path.join(OUTPUT_DIR, "summary_scores.xlsx")

# unknown 类型位置对应的信息会写入下面三个文件：
# 1. missing_txt：只放 replace[i] 的目标字，去重后每 12 字一行
# 2. missing_unknown_tasks.json：详细任务列表，方便追溯来源
# 3. missing_answer.json：尽量兼容你们 action_app.py 使用的 answer.json 结构
MISSING_TXT_FILE = os.path.join(OUTPUT_DIR, "missing_txt")
MISSING_DETAIL_JSON_FILE = os.path.join(OUTPUT_DIR, "missing_unknown_tasks.json")
MISSING_ANSWER_JSON_FILE = os.path.join(OUTPUT_DIR, "missing_answer.json")
MISSING_TXT_LINE_SIZE = 12

# 是否物理清理已有 final 文件里的无效/过期记录
CLEAN_BAD_HISTORY = True


# ==========================================================================
# 2. 通用工具函数
# ==========================================================================

def load_json(filepath: str) -> Any:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"找不到文件：{filepath}")

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    except UnicodeDecodeError:
        with open(filepath, "r", encoding="utf-8-sig") as f:
            return json.load(f)


def sort_key_by_id(item: dict):
    item_id = item.get("id")
    try:
        return (0, int(item_id))
    except Exception:
        return (1, str(item_id))


def read_jsonl(filepath: str) -> List[dict]:
    items = []
    if not os.path.exists(filepath):
        return items

    with open(filepath, "r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                items.append(json.loads(line))
            except json.JSONDecodeError:
                print(f"⚠️ 跳过坏 JSON 行：{filepath} 第 {line_no} 行")

    return items


def rewrite_jsonl(filepath: str, items: List[dict]) -> None:
    items = sorted(items, key=sort_key_by_id)
    with open(filepath, "w", encoding="utf-8") as f:
        for item in items:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")


def append_jsonl_line(file_handle, item: dict) -> None:
    file_handle.write(json.dumps(item, ensure_ascii=False) + "\n")
    file_handle.flush()


def same_clean_source(final_item: dict, clean_item: dict) -> bool:
    """
    防止 clean 文件重新生成后，旧 final 结果被误认为可复用。
    """
    for key in ["phrase", "change", "replace", "model_prediction", "cleaned_prediction"]:
        if final_item.get(key) != clean_item.get(key):
            return False
    return True


# ==========================================================================
# 3. 准确率计算
# ==========================================================================

def calculate_phrase_score(prediction: str, truth: str) -> float:
    if not prediction or not truth:
        return 0.0

    truth_len = len(truth)
    if truth_len == 0:
        return 0.0

    char_weight = 1.0 / truth_len
    score = 0.0

    for p_char, t_char in zip(prediction, truth):
        if p_char == t_char:
            score += char_weight

    return score


def new_action_stats() -> dict:
    return {
        "add": {"total": 0, "correct": 0},
        "replace": {"total": 0, "correct": 0},
        "delete": {"total": 0, "correct": 0},
        "keep": {"total": 0, "correct": 0},
        "unknown": {"total": 0, "correct": 0},
    }


def merge_action_stats(dst: dict, src: dict) -> None:
    for action, stats in src.items():
        if action not in dst:
            dst[action] = {"total": 0, "correct": 0}

        dst[action]["total"] += stats.get("total", 0)
        dst[action]["correct"] += stats.get("correct", 0)



def collect_unknown_action_info(final_items: List[dict], action_dict: dict) -> dict:
    """
    收集 action 判定为 unknown 的位置对应的信息。

    注意：
    - action 类型仍然只由 phrase[i] + change 在 answer_action.json 中决定。
    - 不根据模型预测字符 pred_c 反推动作类型。
    - 即使 cleaned_prediction 是 ERROR，也依然可以收集 unknown 位置；
      因为 unknown 是数据字典缺失问题，和模型是否预测成功无关。

    返回：
    {
        "missing_chars": set(...),              # unknown 位置的 replace[i] 字
        "tasks": [ ... ],                       # 去重后的 base_char/radical/target_char 任务
        "answer_like": {base_char: {radical: target_char}},
        "conflicts": [ ... ]                    # 理论上很少出现；同 base_char+radical 对应多个 target
    }
    """
    missing_chars = set()
    task_map = {}
    answer_like = {}
    conflicts = []

    for item in final_items:
        item_id = item.get("id")
        phrase = item.get("phrase", "")
        truth = item.get("replace", "")
        change = item.get("change", "")
        clean_pred = item.get("cleaned_prediction", "")

        for pos, (orig_c, truth_c) in enumerate(zip(phrase, truth)):
            action = action_dict.get(orig_c, {}).get(change, {}).get("action", "unknown")

            if action != "unknown" or not truth_c:
                continue

            missing_chars.add(truth_c)

            task_key = (orig_c, change, truth_c)
            if task_key not in task_map:
                task_map[task_key] = {
                    "base_char": orig_c,
                    "radical": change,
                    "target_char": truth_c,
                    "count": 0,
                    "example_ids": [],
                    "examples": [],
                }

            task = task_map[task_key]
            task["count"] += 1

            if item_id not in task["example_ids"]:
                task["example_ids"].append(item_id)

            if len(task["examples"]) < 5:
                task["examples"].append({
                    "id": item_id,
                    "phrase": phrase,
                    "replace": truth,
                    "position": pos,
                    "cleaned_prediction": clean_pred,
                })

            # 构造兼容 action_app.py 的 answer.json 结构
            # answer_like[base_char][radical] = target_char
            if orig_c not in answer_like:
                answer_like[orig_c] = {}

            old_target = answer_like[orig_c].get(change)
            if old_target is None:
                answer_like[orig_c][change] = truth_c
            elif old_target != truth_c:
                conflicts.append({
                    "base_char": orig_c,
                    "radical": change,
                    "old_target_char": old_target,
                    "new_target_char": truth_c,
                    "id": item_id,
                    "phrase": phrase,
                    "position": pos,
                })

    tasks = sorted(
        task_map.values(),
        key=lambda x: (x["base_char"], x["radical"], x["target_char"])
    )

    return {
        "missing_chars": missing_chars,
        "tasks": tasks,
        "answer_like": answer_like,
        "conflicts": conflicts,
    }


def write_missing_txt(missing_chars: set, output_file: str = MISSING_TXT_FILE) -> None:
    """
    把 unknown 位置的 replace 字写入 missing_txt。
    默认去重、排序，并按每行 MISSING_TXT_LINE_SIZE 个字输出，方便后续处理。
    """
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    cleaned_chars = sorted({c for c in missing_chars if isinstance(c, str) and c.strip()})

    lines = []
    for i in range(0, len(cleaned_chars), MISSING_TXT_LINE_SIZE):
        lines.append("".join(cleaned_chars[i:i + MISSING_TXT_LINE_SIZE]))

    with open(output_file, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"🧩 unknown 位置 replace 字已写入：{output_file}，共 {len(cleaned_chars)} 个去重字")


def write_missing_json_files(info: dict) -> None:
    """
    写出两个 JSON 文件：
    1. missing_unknown_tasks.json：详细任务信息
    2. missing_answer.json：兼容 action_app.py 的 answer.json 结构

    你后续可以把 action_app.py 里的 ANSWER_FILE 改成：
        ANSWER_FILE = 'eval_results/missing_answer.json'
    或者把 missing_answer.json 复制/合并成 answer.json 后再人工打标。
    """
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    tasks_payload = {
        "description": (
            "These are positions where action was judged as unknown by "
            "answer_action.json lookup using phrase[i] + change. "
            "target_char is replace[i]."
        ),
        "total_unique_tasks": len(info.get("tasks", [])),
        "total_unique_missing_chars": len(info.get("missing_chars", set())),
        "tasks": info.get("tasks", []),
        "conflicts": info.get("conflicts", []),
    }

    with open(MISSING_DETAIL_JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(tasks_payload, f, ensure_ascii=False, indent=4)

    with open(MISSING_ANSWER_JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(info.get("answer_like", {}), f, ensure_ascii=False, indent=4)

    print(f"🧾 unknown 详细任务已写入：{MISSING_DETAIL_JSON_FILE}，共 {len(info.get('tasks', []))} 条去重任务")
    print(f"🧾 前端兼容 answer 结构已写入：{MISSING_ANSWER_JSON_FILE}")

    conflicts = info.get("conflicts", [])
    if conflicts:
        print(f"⚠️ 注意：发现 {len(conflicts)} 条 base_char+radical 对应多个 target_char 的冲突，请查看 missing_unknown_tasks.json 的 conflicts 字段。")


def write_all_missing_outputs(info: dict) -> None:
    write_missing_txt(info.get("missing_chars", set()))
    write_missing_json_files(info)


def merge_missing_infos(infos: List[dict]) -> dict:
    """
    合并多个模型的 unknown 信息。
    因为 unknown 本质上是数据字典缺项，一般各模型会重复发现同一批缺项。
    这里做 union 去重。
    """
    merged_missing_chars = set()
    merged_task_map = {}
    merged_answer_like = {}
    merged_conflicts = []

    for info in infos:
        merged_missing_chars.update(info.get("missing_chars", set()))
        merged_conflicts.extend(info.get("conflicts", []))

        for task in info.get("tasks", []):
            key = (task["base_char"], task["radical"], task["target_char"])

            if key not in merged_task_map:
                merged_task_map[key] = {
                    "base_char": task["base_char"],
                    "radical": task["radical"],
                    "target_char": task["target_char"],
                    "count": 0,
                    "example_ids": [],
                    "examples": [],
                }

            dst = merged_task_map[key]
            dst["count"] += task.get("count", 0)

            for ex_id in task.get("example_ids", []):
                if ex_id not in dst["example_ids"]:
                    dst["example_ids"].append(ex_id)

            for ex in task.get("examples", []):
                if len(dst["examples"]) < 5:
                    dst["examples"].append(ex)

        for base_char, radical_map in info.get("answer_like", {}).items():
            if base_char not in merged_answer_like:
                merged_answer_like[base_char] = {}

            for radical, target_char in radical_map.items():
                old_target = merged_answer_like[base_char].get(radical)
                if old_target is None:
                    merged_answer_like[base_char][radical] = target_char
                elif old_target != target_char:
                    merged_conflicts.append({
                        "base_char": base_char,
                        "radical": radical,
                        "old_target_char": old_target,
                        "new_target_char": target_char,
                        "source": "merge_missing_infos",
                    })

    merged_tasks = sorted(
        merged_task_map.values(),
        key=lambda x: (x["base_char"], x["radical"], x["target_char"])
    )

    return {
        "missing_chars": merged_missing_chars,
        "tasks": merged_tasks,
        "answer_like": merged_answer_like,
        "conflicts": merged_conflicts,
    }


def step3_score_item(item: dict, action_dict: dict) -> Tuple[dict, float, dict]:
    clean_pred = item.get("cleaned_prediction", "")
    truth = item.get("replace", "")
    phrase = item.get("phrase", "")
    change = item.get("change", "")

    score = calculate_phrase_score(clean_pred, truth)
    item["score"] = score

    action_stats = new_action_stats()

    if clean_pred and not clean_pred.startswith("ERROR:"):
        for i, (orig_c, truth_c) in enumerate(zip(phrase, truth)):
            pred_c = clean_pred[i] if i < len(clean_pred) else ""
            action = action_dict.get(orig_c, {}).get(change, {}).get("action", "unknown")

            if action not in action_stats:
                action_stats[action] = {"total": 0, "correct": 0}

            action_stats[action]["total"] += 1

            if pred_c == truth_c:
                action_stats[action]["correct"] += 1

    item["action_stats"] = action_stats

    return item, score, action_stats


# ==========================================================================
# 4. 断点续跑逻辑
# ==========================================================================

def is_valid_final_item(item: dict) -> bool:
    if "score" not in item:
        return False

    if "cleaned_prediction" not in item:
        return False

    if "action_stats" not in item:
        return False

    return True


def load_and_clean_final_history(final_file: str, clean_by_id: Dict[Any, dict]) -> Tuple[List[dict], set, int]:
    history_items = read_jsonl(final_file)
    valid_items = []
    processed_ids = set()
    bad_count = 0

    for item in history_items:
        item_id = item.get("id")

        if item_id is None:
            bad_count += 1
            continue

        clean_item = clean_by_id.get(item_id)
        if clean_item is None:
            # clean 文件中已经没有这个 id，视为过期记录
            bad_count += 1
            continue

        if not same_clean_source(item, clean_item):
            # clean 结果发生变化，旧 final 不能复用
            bad_count += 1
            continue

        if is_valid_final_item(item):
            if item_id not in processed_ids:
                valid_items.append(item)
                processed_ids.add(item_id)
        else:
            bad_count += 1

    valid_items = sorted(valid_items, key=sort_key_by_id)

    if CLEAN_BAD_HISTORY and os.path.exists(final_file):
        rewrite_jsonl(final_file, valid_items)

    return valid_items, processed_ids, bad_count


# ==========================================================================
# 5. 汇总与打印
# ==========================================================================

def summarize_final_items(model_name: str, final_items: List[dict]) -> dict:
    total_evaluated = 0
    total_score = 0.0
    global_action_stats = new_action_stats()

    for item in final_items:
        clean_pred = item.get("cleaned_prediction", "")

        if clean_pred and not clean_pred.startswith("ERROR:"):
            total_evaluated += 1
            total_score += item.get("score", 0.0)
            merge_action_stats(global_action_stats, item.get("action_stats", {}))

    accuracy = total_score / total_evaluated if total_evaluated > 0 else 0.0

    return {
        "model": model_name,
        "total_items": len(final_items),
        "total_evaluated": total_evaluated,
        "total_score": total_score,
        "overall_accuracy": accuracy,
        "global_action_stats": global_action_stats,
    }


def print_model_report(summary: dict) -> None:
    model_name = summary["model"]

    print(f"\n☀️☀️☀️ [{model_name}] 准确率统计完成 ☀️☀️☀️")
    print(f"📊 最终有效统计题目数: {summary['total_evaluated']} / {summary['total_items']}")
    print(f"🏆 词语级综合准确率: {summary['overall_accuracy']:.2%}")

    print("\n🔍 【细粒度动作级统计（按字）】")
    for action in ["keep", "add", "replace", "delete", "unknown"]:
        stats = summary["global_action_stats"].get(action, {"total": 0, "correct": 0})
        if stats["total"] > 0:
            acc = stats["correct"] / stats["total"]
            print(f"  - [{action.upper():<7}] 准确率: {acc:6.2%} (对 {stats['correct']} / 总 {stats['total']} 字)")

    print("=" * 60)


# ==========================================================================
# 6. Excel 汇总
# ==========================================================================

def safe_accuracy(correct: int, total: int):
    if total == 0:
        return None
    return correct / total


def write_excel_summary(summaries: List[dict], output_file: str = EXCEL_SUMMARY_FILE) -> None:
    """
    写出 Excel 汇总表：
    - Summary：模型总分 + 各 action 分项分数，宽表
    - Action_Details：每个模型每个 action 一行，长表
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠️ 未安装 openpyxl，无法写出 Excel。请先运行：pip install openpyxl")
        return

    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    actions = ["keep", "add", "replace", "delete", "unknown"]

    headers = [
        "model",
        "total_items",
        "total_evaluated",
        "overall_accuracy",
    ]

    for action in actions:
        headers.extend([
            f"{action}_accuracy",
            f"{action}_correct",
            f"{action}_total",
        ])

    ws.append(headers)

    for summary in summaries:
        row = [
            summary["model"],
            summary["total_items"],
            summary["total_evaluated"],
            summary["overall_accuracy"],
        ]

        action_stats = summary["global_action_stats"]

        for action in actions:
            stats = action_stats.get(action, {"total": 0, "correct": 0})
            row.extend([
                safe_accuracy(stats["correct"], stats["total"]),
                stats["correct"],
                stats["total"],
            ])

        ws.append(row)

    detail_ws = wb.create_sheet("Action_Details")
    detail_ws.append(["model", "action", "correct", "total", "accuracy"])

    for summary in summaries:
        for action in actions:
            stats = summary["global_action_stats"].get(action, {"total": 0, "correct": 0})
            detail_ws.append([
                summary["model"],
                action,
                stats["correct"],
                stats["total"],
                safe_accuracy(stats["correct"], stats["total"]),
            ])

    # 简单美化
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for sheet in [ws, detail_ws]:
        sheet.freeze_panes = "A2"

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            max_len = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            width = min(max(max_len + 2, 10), 28)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    # 百分比格式
    # Summary: 第4列 overall，以及每个 action 的 accuracy 列
    percent_cols_summary = [4] + [5 + i * 3 for i in range(len(actions))]
    for col_idx in percent_cols_summary:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "0.00%"

    # Action_Details: 第5列
    for row_idx in range(2, detail_ws.max_row + 1):
        detail_ws.cell(row=row_idx, column=5).number_format = "0.00%"

    wb.save(output_file)
    print(f"📘 Excel 汇总表已写出：{output_file}")


# ==========================================================================
# 7. 单模型 / 全模型执行入口
# ==========================================================================

def run_scoring_for_model(model_name: str, write_missing_file: bool = True) -> dict:
    print(f"\n📊 [Phase 3 准确率统计] 当前模型：{model_name}")

    clean_file = os.path.join(OUTPUT_DIR, f"clean_{model_name}.jsonl")
    final_file = os.path.join(OUTPUT_DIR, f"final_{model_name}.jsonl")

    if not os.path.exists(clean_file):
        print(f"⚠️ 找不到 clean 文件：{clean_file}，请先运行 02_clean_text.py。")
        empty_summary = {
            "model": model_name,
            "total_items": 0,
            "total_evaluated": 0,
            "total_score": 0.0,
            "overall_accuracy": 0.0,
            "global_action_stats": new_action_stats(),
            "missing_unknown_info": {
                "missing_chars": set(),
                "tasks": [],
                "answer_like": {},
                "conflicts": [],
            },
        }
        if write_missing_file:
            write_all_missing_outputs(empty_summary["missing_unknown_info"])
        return empty_summary

    action_dict = load_json(ACTION_FILE)

    clean_items = read_jsonl(clean_file)
    clean_items = sorted(clean_items, key=sort_key_by_id)
    rewrite_jsonl(clean_file, clean_items)

    clean_by_id = {item.get("id"): item for item in clean_items}

    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    final_items, processed_ids, bad_count = load_and_clean_final_history(final_file, clean_by_id)

    print(f"📦 clean 记录：{len(clean_items)} 条，已按 id 排序")
    print(f"🧹 历史 final 有效记录：{len(final_items)} 条；清理坏记录/过期记录：{bad_count} 条")

    pending_items = [item for item in clean_items if item.get("id") not in processed_ids]

    if pending_items:
        print(f"⏳ 待评分：{len(pending_items)} / {len(clean_items)} 条")

        new_scored_items = []

        # 评分不调用 API，单线程即可
        for item in tqdm(pending_items, desc=f"评分 {model_name}", unit="条"):
            result_item, _, _ = step3_score_item(item.copy(), action_dict)
            new_scored_items.append(result_item)

        with open(final_file, "a", encoding="utf-8") as fout:
            for item in new_scored_items:
                append_jsonl_line(fout, item)

        final_items.extend(new_scored_items)
    else:
        print(f"✅ {model_name} 已无待评分数据。")

    # 汇总使用完整 final 文件，确保断点续跑后统计的是全量结果
    final_items = read_jsonl(final_file)
    final_items = sorted(final_items, key=sort_key_by_id)
    rewrite_jsonl(final_file, final_items)

    summary = summarize_final_items(model_name, final_items)

    # 收集 action == unknown 的位置对应信息。
    missing_unknown_info = collect_unknown_action_info(final_items, action_dict)
    summary["missing_unknown_info"] = missing_unknown_info

    if write_missing_file:
        write_all_missing_outputs(missing_unknown_info)

    print_model_report(summary)

    return summary


def run_all_scoring() -> None:
    summaries = []
    missing_infos = []

    for model_name in MODEL_LIST:
        summary = run_scoring_for_model(model_name, write_missing_file=False)
        summaries.append(summary)
        missing_infos.append(summary.get("missing_unknown_info", {
            "missing_chars": set(),
            "tasks": [],
            "answer_like": {},
            "conflicts": [],
        }))

    merged_missing_info = merge_missing_infos(missing_infos)
    write_all_missing_outputs(merged_missing_info)
    write_excel_summary(summaries)

    print("\n🎉 Phase 3 全部模型准确率统计完成。")


if __name__ == "__main__":
    run_all_scoring()
