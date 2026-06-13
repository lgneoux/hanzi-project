# ==========================================================================
# Phase 3: 汉字结构测评系统 —— 部首预测评分阶段
# ==========================================================================
# 输入：
# 1. eval_results_radical/clean_radical_{model}.jsonl
# 2. radical_ground_truth.json
#
# 输出：
# 1. eval_results_radical/final_radical_{model}.jsonl
# 2. eval_results_radical/summary_radical_scores.xlsx
# ==========================================================================

import json
import os
from pathlib import Path
from typing import Any, Dict, List, Tuple

from tqdm import tqdm

# ==========================================================================
# 1. 基础配置区
# ==========================================================================

MODEL_LIST = [
    "claude-opus-4-6",
    "gemini-3-flash-preview",
    "glm-4.7",
    "gemini-3-pro-preview",
    "gpt-5.4",
    "deepseek-v3.2",
    "kimi-k2.5"
]

DICT_FILE = "radical_ground_truth.json"
OUTPUT_DIR = "eval_results_radical"
EXCEL_SUMMARY_FILE = os.path.join(OUTPUT_DIR, "summary_radical_scores.xlsx")

# 是否物理清理已有 final 文件里的无效/过期记录
CLEAN_BAD_HISTORY = True

# ==========================================================================
# 2. 通用工具函数 (带编码保护)
# ==========================================================================

def load_json(filepath: str) -> Any:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"找不到文件：{filepath}")

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    except UnicodeDecodeError:
        print(f"⚠️ 检测到编码冲突，尝试使用 GBK 模式读取：{filepath}")
        with open(filepath, "r", encoding="gbk", errors="ignore") as f:
            return json.load(f)

def sort_key_by_id(item: dict):
    item_id = item.get("id")
    try:
        return (0, int(item_id))
    except Exception:
        return (1, str(item_id))

def read_jsonl(filepath: str) -> List[dict]:
    items = []
    if not os.path.exists(filepath):
        return items

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            for line_no, line in enumerate(f, start=1):
                line = line.strip()
                if not line:
                    continue
                try:
                    items.append(json.loads(line))
                except json.JSONDecodeError:
                    print(f"⚠️ 跳过坏 JSON 行：{filepath} 第 {line_no} 行")
    except UnicodeDecodeError:
        print(f"⚠️ 检测到编码冲突，正尝试使用 GBK 兼容模式读取：{filepath}")
        with open(filepath, "r", encoding="gbk", errors="ignore") as f:
            for line_no, line in enumerate(f, start=1):
                line = line.strip()
                if not line:
                    continue
                try:
                    items.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return items

def rewrite_jsonl(filepath: str, items: List[dict]) -> None:
    items = sorted(items, key=sort_key_by_id)
    with open(filepath, "w", encoding="utf-8") as f:
        for item in items:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")

def append_jsonl_line(file_handle, item: dict) -> None:
    file_handle.write(json.dumps(item, ensure_ascii=False) + "\n")
    file_handle.flush()

def same_clean_source(final_item: dict, clean_item: dict) -> bool:
    for key in ["phrase", "model_prediction", "cleaned_prediction"]:
        if final_item.get(key) != clean_item.get(key):
            return False
    return True

# ==========================================================================
# 3. 评分核心逻辑
# ==========================================================================

def step3_score_item(item: dict, radical_dict: dict) -> Tuple[dict, dict]:
    clean_pred = item.get("cleaned_prediction", "")
    phrase = item.get("phrase", "")
    
    # 初始化统计指标
    valid_chars = 0
    correct_chars = 0
    expected_radicals = ""

    is_error = clean_pred.startswith("ERROR:")
    
    for i, char in enumerate(phrase):
        truth_rad = radical_dict.get(char, "未找到")
        
        # 记录理论上的标准答案（方便排查）
        expected_radicals += truth_rad if truth_rad and not truth_rad.startswith("ERROR") else "?"
        
        # 如果爬虫没爬到这个字，或者报错了，我们不计入分母，保证公平
        if not truth_rad or truth_rad == "未找到" or truth_rad.startswith("ERROR"):
            continue
            
        valid_chars += 1
        
        # 防止清洗出来的字符串比 phrase 短导致越界
        pred_c = clean_pred[i] if not is_error and i < len(clean_pred) else ""
        
        if pred_c == truth_rad:
            correct_chars += 1

    item["expected_radicals"] = expected_radicals
    
    if valid_chars > 0:
        item["score"] = correct_chars / valid_chars
    else:
        item["score"] = 0.0
        
    item["char_stats"] = {
        "valid_chars": valid_chars,
        "correct_chars": correct_chars
    }
    
    return item

# ==========================================================================
# 4. 断点续跑与数据加载
# ==========================================================================

def load_and_clean_final_history(final_file: str, clean_by_id: Dict[Any, dict]) -> Tuple[List[dict], set, int]:
    history_items = read_jsonl(final_file)
    valid_items = []
    processed_ids = set()
    bad_count = 0

    for item in history_items:
        item_id = item.get("id")

        if item_id is None or item_id not in clean_by_id:
            bad_count += 1
            continue

        clean_item = clean_by_id[item_id]

        if not same_clean_source(item, clean_item) or "score" not in item:
            bad_count += 1
            continue

        if item_id not in processed_ids:
            valid_items.append(item)
            processed_ids.add(item_id)

    valid_items = sorted(valid_items, key=sort_key_by_id)

    if CLEAN_BAD_HISTORY and os.path.exists(final_file):
        rewrite_jsonl(final_file, valid_items)

    return valid_items, processed_ids, bad_count

# ==========================================================================
# 5. Excel 汇总与打印
# ==========================================================================

def summarize_final_items(model_name: str, final_items: List[dict]) -> dict:
    total_items = len(final_items)
    word_exact_match = 0
    
    total_valid_chars = 0
    total_correct_chars = 0

    for item in final_items:
        stats = item.get("char_stats", {})
        valid = stats.get("valid_chars", 0)
        correct = stats.get("correct_chars", 0)
        
        total_valid_chars += valid
        total_correct_chars += correct
        
        # 如果一个词里面所有有效的字都预测对了，并且至少包含一个有效字，算作词语级全对
        if valid > 0 and valid == correct:
            word_exact_match += 1

    char_accuracy = total_correct_chars / total_valid_chars if total_valid_chars > 0 else 0.0
    word_accuracy = word_exact_match / total_items if total_items > 0 else 0.0

    return {
        "model": model_name,
        "total_items": total_items,
        "word_exact_match": word_exact_match,
        "word_accuracy": word_accuracy,
        "total_valid_chars": total_valid_chars,
        "total_correct_chars": total_correct_chars,
        "char_accuracy": char_accuracy,
    }

def print_model_report(summary: dict) -> None:
    print(f"\n☀️☀️☀️ [{summary['model']}] 准确率统计完成 ☀️☀️☀️")
    print(f"📊 总测评词语数: {summary['total_items']}")
    print(f"🎯 词语级完全准确率: {summary['word_accuracy']:.2%} ({summary['word_exact_match']} / {summary['total_items']})")
    print(f"🎯 单字级部首准确率: {summary['char_accuracy']:.2%} ({summary['total_correct_chars']} / {summary['total_valid_chars']})")
    print("=" * 60)

def write_excel_summary(summaries: List[dict], output_file: str = EXCEL_SUMMARY_FILE) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠️ 未安装 openpyxl，无法写出 Excel。请先运行：pip install openpyxl")
        return

    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Radical_Summary"

    headers = [
        "模型 (Model)",
        "总词语数 (Total Words)",
        "词语全对数 (Exact Words)",
        "词语级准确率 (Word Acc)",
        "总有效汉字数 (Valid Chars)",
        "部首正确数 (Correct Rads)",
        "单字级准确率 (Char Acc)",
    ]
    ws.append(headers)

    for summary in summaries:
        row = [
            summary["model"],
            summary["total_items"],
            summary["word_exact_match"],
            summary["word_accuracy"],
            summary["total_valid_chars"],
            summary["total_correct_chars"],
            summary["char_accuracy"],
        ]
        ws.append(row)

    # 简单美化
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="center")

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            # 中文字符宽一点，稍微多加点宽度
            max_len = max(max_len, len(value) * 1.5)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 30)

    # 百分比格式化
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=4).number_format = "0.00%" # 词语级准确率
        ws.cell(row=row_idx, column=7).number_format = "0.00%" # 单字级准确率

    wb.save(output_file)
    print(f"📘 Excel 汇总表已写出：{output_file}")

# ==========================================================================
# 6. 执行入口
# ==========================================================================

def run_scoring_for_model(model_name: str, radical_dict: dict) -> dict:
    print(f"\n📊 [Phase 3 部首评分] 当前模型：{model_name}")

    clean_file = os.path.join(OUTPUT_DIR, f"clean_radical_{model_name}.jsonl")
    final_file = os.path.join(OUTPUT_DIR, f"final_radical_{model_name}.jsonl")

    if not os.path.exists(clean_file):
        print(f"⚠️ 找不到 clean 文件：{clean_file}，跳过评分。")
        return {
            "model": model_name,
            "total_items": 0, "word_exact_match": 0, "word_accuracy": 0.0,
            "total_valid_chars": 0, "total_correct_chars": 0, "char_accuracy": 0.0,
        }

    clean_items = read_jsonl(clean_file)
    clean_items = sorted(clean_items, key=sort_key_by_id)
    rewrite_jsonl(clean_file, clean_items)

    clean_by_id = {item.get("id"): item for item in clean_items}
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    final_items, processed_ids, bad_count = load_and_clean_final_history(final_file, clean_by_id)

    print(f"📦 clean 记录：{len(clean_items)} 条")
    print(f"🧹 历史 final 有效记录：{len(final_items)} 条；清理过期记录：{bad_count} 条")

    pending_items = [item for item in clean_items if item.get("id") not in processed_ids]

    if pending_items:
        print(f"⏳ 待评分：{len(pending_items)} / {len(clean_items)} 条")
        new_scored_items = []

        for item in tqdm(pending_items, desc=f"评分 {model_name}", unit="条"):
            result_item = step3_score_item(item.copy(), radical_dict)
            new_scored_items.append(result_item)

        with open(final_file, "a", encoding="utf-8") as fout:
            for item in new_scored_items:
                append_jsonl_line(fout, item)

        final_items.extend(new_scored_items)
    else:
        print(f"✅ {model_name} 已无待评分数据。")

    final_items = read_jsonl(final_file)
    final_items = sorted(final_items, key=sort_key_by_id)
    rewrite_jsonl(final_file, final_items)

    summary = summarize_final_items(model_name, final_items)
    print_model_report(summary)

    return summary

def run_all_scoring() -> None:
    print("📖 正在加载部首标准答案字典...")
    radical_dict = load_json(DICT_FILE)
    print(f"✅ 成功加载 {len(radical_dict)} 个汉字的部首信息。")

    summaries = []
    for model_name in MODEL_LIST:
        summary = run_scoring_for_model(model_name, radical_dict)
        summaries.append(summary)

    write_excel_summary(summaries)
    print("\n🎉 Phase 3 全部模型部首准确率统计完成。")

if __name__ == "__main__":
    run_all_scoring()